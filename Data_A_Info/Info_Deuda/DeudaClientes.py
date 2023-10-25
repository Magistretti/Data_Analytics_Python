import os
import math
import re
import numpy as np
from DatosLogin import login
from PIL import Image
from InfoDeuda.Dias_Deuda import dias_Deuda
import pandas as pd
import dataframe_image as dfi
import pyodbc #Library to connect to Microsoft SQL Server
from DatosLogin import login #Holds connection data to the SQL Server
import sys
import pathlib
from calendar import monthrange
from datetime import datetime
from datetime import timedelta
import datetime
sys.path.insert(0,str(pathlib.Path(__file__).parent.parent))
import logging
import matplotlib.pyplot as plt
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
        , level=logging.INFO
)
logger = logging.getLogger(__name__)

#################################

tiempoInicio = pd.to_datetime("today")
#########
# Formating display of dataframes with comma separator for numbers
pd.options.display.float_format = "{:20,.0f}".format 
#########

try:
    db_conex = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};\
        SERVER="+login[0]+";\
        DATABASE="+login[1]+";\
        UID="+login[2]+";\
        PWD="+ login[3]
    )
except Exception as e:
    listaErrores = e.args[1].split(".")
    logger.error("\nOcurrió un error al conectar a SQL Server: ")
    for i in listaErrores:
        logger.error(i)
    exit()

###########################################################################
##############################   Deuda Clientes ######################################
###########################################################################
deuda = pd.read_sql(
 ''' 
    SELECT
            CAST(FRD.[NROCLIENTE] as VARCHAR) as 'NROCLIENTE'
            ,RTRIM(Cli.NOMBRE) as 'NOMBRE'

            ,MIN(CAST(FRD.[FECHASQL] as date)) as 'FECHA_1erRemito'
            ,MAX(CAST(FRD.[FECHASQL] as date)) as 'FECHA_UltRemito'

            ----Días Entre Remitos
            ,IIF(MIN(CAST(FRD.[FECHASQL] as date)) = MAX(CAST(FRD.[FECHASQL] as date))
            	, -1
            	, DATEDIFF(DAY,MAX(CAST(FRD.[FECHASQL] as date)),MIN(CAST(FRD.[FECHASQL] as date)))
            ) as 'Días Entre Remitos'

            --,sum(FRD.[IMPORTE]) as 'ConsumoHistorico'

            ----Consumo Diario
            --,sum(FRD.[IMPORTE])/IIF(MIN(CAST(FRD.[FECHASQL] as date)) = MAX(CAST(FRD.[FECHASQL] as date))
            --	, -1
            --	, DATEDIFF(DAY,MAX(CAST(FRD.[FECHASQL] as date)),MIN(CAST(FRD.[FECHASQL] as date)))
            --) as 'Consumo Diario'

            ,CAST(ROUND(MIN(Cli.SALDOPREPAGO - Cli.SALDOREMIPENDFACTU),0) as int) as 'SALDOCUENTA'

            , DATEDIFF(DAY, MAX(CAST(FRD.[FECHASQL] as date)), CAST(GETDATE() as date)) as 'Días Desde Última Compra'

            , ISNULL(RTRIM(Vend.NOMBREVEND),'') as 'Vendedor'
			, cli.LIMITECREDITO as 'Acuerdo de Descubierto $'
			, sum(FRD.importe) as 'venta en cta/cte'
            FROM [Rumaos].[dbo].[FacRemDet] as FRD with (NOLOCK)
            INNER JOIN dbo.FacCli as Cli with (NOLOCK)
                ON FRD.NROCLIENTE = Cli.NROCLIPRO
            LEFT OUTER JOIN dbo.Vendedores as Vend with (NOLOCK)
	            ON Cli.NROVEND = Vend.NROVEND

            where FRD.NROCLIENTE > '100000'
                AND (Cli.SALDOPREPAGO - Cli.SALDOREMIPENDFACTU) < -10000
                and Cli.ListaSaldoCC = 1
                and FECHASQL >= '20140101' and FECHASQL <= CAST(GETDATE() as date)

            group by FRD.NROCLIENTE, Cli.NOMBRE, Vend.NOMBREVEND,cli.LIMITECREDITO
            order by Cli.NOMBRE
 ''' 
  ,db_conex)


deuda = deuda.convert_dtypes()
deuda['NROCLIENTE'] = deuda['NROCLIENTE'].astype(int)
deuda['Cupo Disponible']= deuda['SALDOCUENTA']+deuda['Acuerdo de Descubierto $']
deuda=deuda.merge(dias_Deuda,on=['NROCLIENTE'],how='outer')


def categorizar_deuda(dias_adeuda, dias_ultima):
    if dias_adeuda < 20:
        return '1-Normal'
    elif dias_adeuda < 40:
        return '2-Moroso'
    elif dias_adeuda <= 60:
        return '3-Moroso Grave'
    elif dias_adeuda > 60:
        return '4-Gestión Judicial'

deuda['Tipo de deudor'] = deuda.apply(lambda x: categorizar_deuda(x['Dias de Venta Adeudado'], x['Días Desde Última Compra']), axis=1)

# Función para rellenar los valores faltantes según el tipo de dato
deuda['NOMBRE'] = deuda['NOMBRE'].fillna('')
deuda['Cupo Disponible'] = deuda['Cupo Disponible'].fillna(0)
deuda['Acuerdo'] = deuda['Cupo Disponible'].apply(lambda x: 'Deudor Excedido' if x < 0 else 'Deudor No Excedido')

def calcular_interes(row):
    if row['Dias de Venta Adeudado'] > 15:
        interes_diario = 1.0 / 365  # Tasa de interés diaria (100% anual)
        dias_adeudados = row['Dias de Venta Adeudado'] - 15  # Días de venta adeudados que exceden los 20 días
        saldo_adeudado = -row['SALDOCUENTA']  # Saldo adeudado (en valor absoluto)
        interes_cobrado = saldo_adeudado * interes_diario * dias_adeudados  # Interés cobrado
        return interes_cobrado
    else:
        return 0

# Aplicar la función a cada fila del DataFrame para crear la nueva columna
deuda['Interes por Mora'] = deuda.apply(calcular_interes, axis=1)

        
deuda = deuda.reindex(columns=['NROCLIENTE','NOMBRE','SALDOCUENTA','Cantidad Adeudada','Días Desde Última Compra','Acuerdo de Descubierto $','Dias de Venta Adeudado','Cupo Disponible','Interes por Mora','Vendedor','Tipo de deudor','Acuerdo'])
deuda = deuda.loc[deuda["NOMBRE"] != " ",:]
deuda = deuda.loc[deuda["NOMBRE"] != "",:]
deuda.sort_values(['NOMBRE'])
normal= deuda.loc[deuda["Tipo de deudor"] == "1-Normal",:]
moroso= deuda.loc[deuda["Tipo de deudor"] == "2-Moroso",:]
morosoG= deuda.loc[deuda["Tipo de deudor"] == "3-Moroso Grave",:]
gestionJ= deuda.loc[deuda["Tipo de deudor"] == "4-Gestión Judicial",:]
normal=normal.sort_values(['NOMBRE'])

color_map = {
    '1-Normal': 'green',
    '2-Moroso': 'orange',
    '3-Moroso Grave': 'red',
    '4-Gestión Judicial': 'maroon',
    'TOTAL':'black',
    ' ':'black'
}
color_map2 = {
    'Deudor No Excedido': 'green',
    'Deudor Excedido': 'red',
    ' ':'black'

}


total = deuda.reindex(columns=['SALDOCUENTA','Tipo de deudor'])
total = total.groupby(
        ["Tipo de deudor"]
        , as_index=False
    ).sum()

total['Participacion']=total['SALDOCUENTA']/total['SALDOCUENTA'].sum()

total.loc["colTOTAL"]= pd.Series(
    total.sum()
    , index=["SALDOCUENTA"]
)
total.fillna({"Tipo de deudor":"TOTAL"}, inplace=True)
total.fillna({"Participacion": total['SALDOCUENTA'].sum()/total['SALDOCUENTA'].sum()}, inplace=True)

ventas15Dias = pd.read_sql(
 ''' 
    DECLARE @ayer DATETIME
	SET @ayer = DATEADD(day, -1, CAST(GETDATE() AS date))
	DECLARE @inicioMesActual DATETIME
	SET @inicioMesActual = DATEADD(month, DATEDIFF(month, 0, @ayer), 0)

	DECLARE @inicioMesAnterior DATETIME
	SET @inicioMesAnterior = DATEADD(M,-1,@inicioMesActual)

	--Divide por la cant de días del mes anterior y multiplica por la cant de días del
	--mes actual
	
	DECLARE @hoy DATETIME
	SET @hoy = DATEADD(DAY, DATEDIFF(DAY, 0, CURRENT_TIMESTAMP), 0)

	select NROCLIPRO as NROCLIENTE,sum(IMPTOTAL) as 'Ventas Ultimos 15 Dias' from FacRemGen
	where FECHASQL <= @hoy
	and FECHASQL >= DATEADD(DAY,-15,CAST(GETDATE() AS date))
	and NROCLIPRO >= '100000'
	group by NROCLIPRO
 ''' 
  ,db_conex)
ventas15Dias=ventas15Dias.convert_dtypes()
deuda = deuda.merge(ventas15Dias,on='NROCLIENTE',how='left')



deuda=deuda.sort_values('SALDOCUENTA', ascending=True)
deuda.loc["colTOTAL"]= pd.Series(
    deuda.sum(numeric_only=True)
    , index=["SALDOCUENTA","Ventas Ultimos 15 Dias"]
)
deuda.fillna({"NOMBRE":"TOTAL"}, inplace=True)
deuda['Tipo de deudor']=deuda['Tipo de deudor'].fillna(' ')
deuda['Acuerdo']=deuda['Acuerdo'].fillna(' ')

deuda = deuda.reindex(columns=['NROCLIENTE','NOMBRE','SALDOCUENTA','Cantidad Adeudada','Dias de Venta Adeudado','Días Desde Última Compra','Ventas Ultimos 15 Dias','Acuerdo de Descubierto $','Cupo Disponible','Interes por Mora','Vendedor','Tipo de deudor','Acuerdo'])


def _estiladorVtaTituloD(df,list_Col_Perc,list_Col_Numpes, list_Col_Num,listaporcentaje, titulo):
    """
This function will return a styled dataframe that must be assign to a variable.
ARGS:
    df: Dataframe that will be styled.
    list_Col_Num: List of numeric columns that will be formatted with
    zero decimals and thousand separator.
    list_Col_Perc: List of numeric columns that will be formatted 
    as percentage.
    titulo: String for the table caption.
    """
    excluded_rows = df.index[-1:]
    resultado = df.style \
        .format("$ {0:,.0f}", subset=list_Col_Numpes) \
        .format("{0:,.0f}", subset=list_Col_Num) \
        .format("{:,.2%}", subset=listaporcentaje) \
        .hide_index() \
        .set_caption(
            titulo
            + "<br>"
            + ((tiempoInicio-pd.to_timedelta(1,"days")).strftime("%d/%m/%y"))
            + "<br>") \
        .set_properties(subset= list_Col_Perc + list_Col_Num +list_Col_Numpes+listaporcentaje
            , **{"text-align": "center", "width": "100px"}) \
        .set_properties(border= "2px solid black") \
        .set_table_styles([
            {"selector": "caption", 
                "props": [
                    ("font-size", "20px")
                    ,("text-align", "center")
                ]
            }
            , {"selector": "th", 
                "props": [
                    ("text-align", "center")
                    ,("background-color","black")
                    ,("color","white")
                ]
            }
        ]) \
        .apply(lambda x: ["background: black" if x.name == "colTOTAL" 
            else "" for i in x]
            , axis=1) \
        .apply(lambda x: ["color: white" if x.name == "colTOTAL" 
            else "" for i in x]
            , axis=1)\
        .applymap(lambda x: f'background-color: {color_map[x]}' if x not in excluded_rows else '', subset=['Tipo de deudor']) \
        .applymap(lambda x: f'background-color: {color_map2[x]}' if x not in excluded_rows else '', subset=['Acuerdo'])
    return resultado


######### LE DOY FORMATO AL DATAFRAME
def _estiladorVtaTituloDTotal(df,list_Col_Numpes,listaporcentaje, titulo):
    """
This function will return a styled dataframe that must be assign to a variable.
ARGS:
    df: Dataframe that will be styled.
    list_Col_Num: List of numeric columns that will be formatted with
    zero decimals and thousand separator.
    list_Col_Perc: List of numeric columns that will be formatted 
    as percentage.
    titulo: String for the table caption.
    """
    resultado = df.style \
        .format("$ {0:,.0f}", subset=list_Col_Numpes) \
        .format("{:,.2%}", subset=listaporcentaje) \
        .hide_index() \
        .set_caption(
            titulo
            + "<br>"
            + ((tiempoInicio-pd.to_timedelta(1,"days")).strftime("%d/%m/%y"))
            + "<br>") \
        .set_properties(subset= list_Col_Numpes+listaporcentaje
            , **{"text-align": "center", "width": "100px"}) \
        .set_properties(border= "2px solid black") \
        .set_table_styles([
            {"selector": "caption", 
                "props": [
                    ("font-size", "20px")
                    ,("text-align", "center")
                ]
            }
            , {"selector": "th", 
                "props": [
                    ("text-align", "center")
                    ,("background-color","black")
                    ,("color","white")
                ]
            }
        ]) \
        .applymap(lambda x: f'background-color: {color_map[x]}', subset=['Tipo de deudor'])\
        .apply(lambda x: ["background: black" if x.name == "colTOTAL" 
            else "" for i in x]
            , axis=1) \
        .apply(lambda x: ["color: white" if x.name == "colTOTAL" 
            else "" for i in x]
            , axis=1)
     
    return resultado


#  columnas sin decimales
numCols1=[]
numCols = [ 'Días Desde Última Compra','Dias de Venta Adeudado','Cantidad Adeudada']
colpesos=[ 'SALDOCUENTA','Acuerdo de Descubierto $','Cupo Disponible','Interes por Mora']
coldeuda=['SALDOCUENTA']
# Columnas Porcentaje
percColsstr = ['Tipo de deudor','Acuerdo']
percColsstr1 = ['Tipo de deudor']
percCol = ['Participacion']



deuda = _estiladorVtaTituloD(deuda,percColsstr,colpesos,numCols,numCols1, "deuda")
moroso = _estiladorVtaTituloD(moroso,percColsstr,colpesos,numCols,numCols1, "deuda")

morosoG = _estiladorVtaTituloD(morosoG,percColsstr,colpesos,numCols,numCols1, "deuda")

gestionJ = _estiladorVtaTituloD(gestionJ,percColsstr,colpesos,numCols,numCols1, "deuda")


total = _estiladorVtaTituloDTotal(total,coldeuda,percCol, "deuda")


### APLICO EL FORMATO A LA TABLA
ubicacion = "C:/Informes/InfoDeuda/"
nombreN = "Deudores_Comerciales.png"
nombreM= "Deudores_Morosos.png"
nombreMG = "Deudores_MorososG.png"
nombreGJ = "Deudores_GestionJ.png"
nombreT='Deuda_Comercial.png'

def df_to_image(df, ubicacion, nombre):
    """
    Esta función usa las biblioteca "dataframe_Image as dfi" y "os" para 
    generar un archivo .png de un dataframe. Si el archivo ya existe, este será
    reemplazado por el nuevo archivo.

    Args:
        df: dataframe a convertir
         ubicacion: ubicacion local donde se quiere grabar el archivo
          nombre: nombre del archivo incluyendo extensión .png (ej: "hello.png")

    """
    if os.path.exists(ubicacion+nombre):
        os.remove(ubicacion+nombre)
        dfi.export(df, ubicacion+nombre,max_rows=-1)
    else:
        dfi.export(df, ubicacion+nombre,max_rows=-1)

df_to_image(deuda, ubicacion, nombreN)
df_to_image(moroso, ubicacion, nombreM)
df_to_image(morosoG, ubicacion, nombreMG)
df_to_image(gestionJ, ubicacion, nombreGJ)
df_to_image(total, ubicacion, nombreT)



imgRGBA = Image.open(ubicacion+nombreN)
imRGB = Image.new('RGB', imgRGBA.size, (255, 255, 255))
imRGB.paste(imgRGBA, mask=imgRGBA.split()[3])

nombrePDF = "Deudores_Comerciales.pdf"

# Saving has a PDF and avoiding permission error
if os.path.exists(ubicacion + nombrePDF):
    os.remove(ubicacion + nombrePDF)
    imRGB.save(ubicacion + nombrePDF, "PDF", resolution=90.0, save_all=True)
else:
    imRGB.save(ubicacion + nombrePDF, "PDF", resolution=90.0, save_all=True)

### EXCEL
ubicacionExcel = "C:/Informes/InfoDeuda/"

nombreExcel = "Clientes_Deudores.xlsx"

### IMPRIMO LA Excel 

def df_to_Excel(df, ubicacion, nombre):
    """
    Esta función usa las biblioteca "dataframe_Image as dfi" y "os" para 
    generar un archivo .png de un dataframe. Si el archivo ya existe, este será
    reemplazado por el nuevo archivo.

    Args:
        df: dataframe a convertir
         ubicacion: ubicacion local donde se quiere grabar el archivo
          nombre: nombre del archivo incluyendo extensión .png (ej: "hello.png")

    """
    if os.path.exists(ubicacion+nombre):
        os.remove(ubicacion+nombre)
        df.to_excel(ubicacion+nombre)
    else:
        df.to_excel(ubicacion+nombre)

df_to_Excel(deuda, ubicacionExcel, nombreExcel)


import openpyxl
from openpyxl.styles import PatternFill

# Cargar el archivo Excel
archivo_excel = openpyxl.load_workbook('C:/Informes/InfoDeuda/Clientes_Deudores.xlsx')

# Obtener la hoja de cálculo activa
hoja_activa = archivo_excel.active

# Agrandar las columnas A a H
hoja_activa.column_dimensions.group('B', 'L', hidden=False)

# Obtener el número de la última fila escrita
ultima_fila = hoja_activa.max_row

# Pintar la última fila desde la columna B hasta la K
fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
for col in range(2, 12):  # Columnas B a K
    celda = hoja_activa.cell(row=ultima_fila, column=col)
    celda.fill = fill

# Guardar los cambios en el archivo Excel
archivo_excel.save('C:/Informes/InfoDeuda/Clientes_Deudores.xlsx')

# Cerrar el archivo
archivo_excel.close()

#############
# Timer
tiempoFinal = pd.to_datetime("today")
logger.info(
    "\nInfo Volumen de Ventas"
    + "\nTiempo de Ejecucion Total: "
    + str(tiempoFinal-tiempoInicio)
)










